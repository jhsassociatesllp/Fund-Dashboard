"""
Parses admin-uploaded Client Master / Corpus Movement files (CSV or Excel)
into the document shape stored in MongoDB.

Column headers are matched loosely - lowercased with punctuation and spaces
stripped - against a set of known aliases, so small variations in an
uploaded template (extra spaces, "Email" vs "EmailID", etc.) still map to
the right field instead of being silently dropped.

PDF/Word are intentionally not parsed here: reliably extracting a table
from an arbitrary PDF or Word layout is not something that can be done
correctly in general, and a misread cell in this data (a bank account
number, a commitment amount) is a real-money mistake. Admins uploading
those should export to CSV/Excel first.
"""

import re
from io import BytesIO

import pandas as pd


def _normalize_header(header: str) -> str:
    return re.sub(r"[^a-z0-9]", "", str(header).strip().lower())


def read_table(filename: str, content: bytes) -> pd.DataFrame:
    lower = (filename or "").lower()

    if lower.endswith(".csv"):
        return pd.read_csv(BytesIO(content), dtype=str, keep_default_na=False)
    if lower.endswith((".xlsx", ".xlsm")):
        return pd.read_excel(BytesIO(content), dtype=str, engine="openpyxl", keep_default_na=False)
    if lower.endswith(".xls"):
        return pd.read_excel(BytesIO(content), dtype=str, engine="xlrd", keep_default_na=False)
    if lower.endswith((".pdf", ".doc", ".docx")):
        raise ValueError(
            "PDF and Word files aren't supported for data upload - extracting tables from them "
            "reliably enough for financial data (bank details, commitment amounts) isn't possible. "
            "Please save the sheet as CSV or Excel (.xlsx) and upload that instead."
        )
    raise ValueError(f"Unsupported file type: '{filename}'. Please upload a CSV or Excel (.xlsx/.xls) file.")


def _rows_from_dataframe(df: pd.DataFrame) -> list[dict]:
    normalized_columns = {col: _normalize_header(col) for col in df.columns}
    rows = []
    for _, series in df.iterrows():
        row = {normalized_columns[col]: series[col] for col in df.columns}
        if any(str(value).strip() for value in row.values()):
            rows.append(row)
    return rows


def _text(row: dict, *keys: str, default: str = "-") -> str:
    for key in keys:
        value = row.get(key)
        if value is not None and str(value).strip():
            return str(value).strip()
    return default


def _number(row: dict, *keys: str) -> float:
    for key in keys:
        value = row.get(key)
        if value is None:
            continue
        raw = str(value).strip()
        if not raw or raw == "-":
            continue
        cleaned = re.sub(r"[^0-9.\-]", "", raw)
        if cleaned in ("", "-", ".", "-."):
            continue
        try:
            return float(cleaned)
        except ValueError:
            continue
    return 0.0


def parse_client_master(df: pd.DataFrame) -> tuple[list[dict], list[str]]:
    rows = _rows_from_dataframe(df)
    docs: list[dict] = []
    warnings: list[str] = []

    for i, row in enumerate(rows, start=2):
        investor_name = _text(row, "investorname", default="")
        client_code = _text(row, "clientid", "clientcode", default="")
        if not investor_name or not client_code:
            warnings.append(f"Row {i}: skipped - missing Investor Name or Client ID")
            continue

        docs.append({
            "investor_name": investor_name,
            "client_class": _text(row, "class"),
            "management_fees": _text(row, "managementfees"),
            "client_code": client_code,
            "dp_id": _text(row, "dpid"),
            "im_signing_date": _text(row, "imsigningdate"),
            "status": _text(row, "statusindividualcompany", "status"),
            "nominee_1_name": _text(row, "nominee1name"),
            "dob_or_incorporation_date": _text(
                row, "dateofbirthdateofincorporation", "dob", "dateofincorporation", "dateofbirth"
            ),
            "joint_holder_name": _text(row, "jointholdername"),
            "mobile_no": _text(row, "mobileno", "mobilenumber"),
            "email_id": _text(row, "emailid", "email"),
            "address_1": _text(row, "address1", "address"),
            "city": _text(row, "city"),
            "pin_code": _text(row, "pincode"),
            "country": _text(row, "country"),
            "bank_name": _text(row, "bankname"),
            "bank_account_no": _text(row, "bankaccountno", "bankaccountnumber"),
            "bank_account_type": _text(row, "bankaccounttype"),
            "bank_ifsc_code": _text(row, "bankifsccode", "ifsccode", "ifsc"),
            "commitment_amount": _number(row, "commitmentamount"),
            "top_up_amount": _number(row, "topupamount"),
            "commitment_reduced": _number(row, "commitmentreduced"),
            "total_commitment": _number(row, "totalcommitment"),
            "initial_contribution": _number(row, "intialcontribution", "initialcontribution"),
            "distributor_name": _text(row, "distributorname"),
            "distributor_code": _text(row, "distributorcode"),
            "side_letters": _text(row, "sideletters"),
            "remarks": _text(row, "remarks"),
            "scheme": _text(row, "scheme"),
        })

    return docs, warnings


def parse_generic(df: pd.DataFrame) -> tuple[list[dict], list[str]]:
    """Parses an upload for a NAV Income/Expense category whose columns aren't
    standardized yet - each row is kept as-is (original header -> value), with
    fully blank rows dropped.
    """
    docs: list[dict] = []
    for _, series in df.iterrows():
        row = {str(col).strip(): str(series[col]).strip() for col in df.columns}
        if any(value for value in row.values()):
            docs.append({"data": row})
    return docs, []


def _to_float_or_none(value) -> float | None:
    if value is None:
        return None
    raw = str(value).strip()
    if not raw:
        return None
    cleaned = re.sub(r"[^0-9.\-]", "", raw)
    if cleaned in ("", "-", ".", "-."):
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


DIFF_TOLERANCE = 0.01


def _is_diff_header(text) -> bool:
    """True for a header that marks a Diff column - "Diff", "Difference", or a
    disambiguated variant like "Diff (Class)" (Corpus Out repeats "Diff" several times
    in its auditor block, so some sheets suffix it to tell the columns apart)."""
    return _normalize_header(text).startswith("diff")


def _parse_sheet_with_auditor_validation(
    filename: str,
    content: bytes,
    category_label: str,
    field_aliases: dict[str, str] | None = None,
) -> tuple[list[dict], list[str]]:
    """Parses an upload that cross-checks the fund's figures against an auditor's
    recalculation when the sheet has one.

    These sheets use a two-row header: row 2 holds the real field names
    (Instrument, ISIN, Trade Qty, ...); row 1 is normally blank except above a
    block of columns the auditor added to re-verify specific fields - a merged
    label (e.g. "Validation by auditor") sits above that block, which is laid
    out as repeating (field, Diff) pairs, e.g. "Trade Qty", "Diff", "Trade Rate",
    "Diff", ... A non-zero Diff means the auditor's recalculation doesn't match
    the fund's original figure for that field on that row. Row 1 can carry more than
    one label (e.g. Corpus In's "In" over the base columns plus "As per Validator" over
    the recheck block, or Corpus Out's "Capital Amount Working"/"Bank Check"/"EXIT LOAD"
    group titles alongside its own "As per Validator" block) - only a label that reads
    as an auditor/validator recheck is treated as the validation block; other row-1
    group labels just describe a block of base columns and stay in base_cols.

    An auditor column is usually named exactly like the base field it re-checks
    (e.g. "Trade Qty" on both sides). Where a sheet uses shorter names in the
    auditor block (e.g. "Quantity" re-checking "Holding Qty"), pass
    `field_aliases` mapping the normalized auditor name to the normalized base
    field name.

    Only .xlsx/.xlsm carry merged cells / multi-row headers, so CSV and legacy
    .xls uploads fall back to the flat parse_generic() read (no validation -
    status/errors are simply absent on those rows). A sheet with no row-1 merge at
    all (no auditor block, just a plain single header row) also falls back to a flat
    read the same way, rather than assuming a two-row header that isn't there.
    """
    field_aliases = field_aliases or {}
    lower = (filename or "").lower()
    if not lower.endswith((".xlsx", ".xlsm")):
        df = read_table(filename, content)
        return parse_generic(df)

    try:
        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(content), data_only=True)
        ws = wb.active
    except Exception as exc:
        raise ValueError(f"Could not read '{filename}' as an Excel file: {exc}")

    max_col = ws.max_column
    max_row = ws.max_row

    has_row1_merge = any(r.min_row <= 1 <= r.max_row for r in ws.merged_cells.ranges)
    if not has_row1_merge:
        if max_row < 2:
            raise ValueError(f"No data rows found in the uploaded file for {category_label}.")
        headers = [str(ws.cell(row=1, column=c).value or "").strip() for c in range(1, max_col + 1)]
        docs: list[dict] = []
        for r in range(2, max_row + 1):
            raw_values = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
            if not any(str(v).strip() for v in raw_values if v is not None):
                continue
            row: dict[str, str] = {}
            for i, header in enumerate(headers):
                if not header:
                    continue
                row[header] = "" if raw_values[i] is None else str(raw_values[i]).strip()
            if row:
                docs.append({"data": row, "status": None, "errors": []})
        return docs, []

    if max_row < 3:
        raise ValueError(f"No data rows found in the uploaded file for {category_label}.")

    # Expand row-1 labels across any merged ranges so a merged cell like
    # "Validation by auditor" marks every column it visually spans, not just
    # its top-left cell (the only one that carries a value once read back).
    row1 = [ws.cell(row=1, column=c).value for c in range(1, max_col + 1)]
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row <= 1 <= merged_range.max_row:
            label = ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
            for c in range(merged_range.min_col, merged_range.max_col + 1):
                row1[c - 1] = label

    # Some sheets type the "Validation by auditor" label into a single cell without an
    # actual Excel merge (merged_cells.ranges is empty for them) - it still visually reads
    # as spanning the block since the following header cells are left blank. The auditor
    # block is always the tail of the sheet, so once a label is seen, forward-fill it to
    # the last column instead of only trusting real merges.
    last_label = None
    for c in range(max_col):
        if row1[c] not in (None, ""):
            last_label = row1[c]
        elif last_label is not None:
            row1[c] = last_label

    row2 = [str(ws.cell(row=2, column=c).value or "").strip() for c in range(1, max_col + 1)]

    auditor_cols = [i for i, v in enumerate(row1) if str(v or "").strip() and re.search(r"valid|audit", str(v), re.I)]
    auditor_col_set = set(auditor_cols)
    base_cols = [i for i in range(max_col) if i not in auditor_col_set]

    base_index_by_name = {}
    for i in base_cols:
        name = _normalize_header(row2[i])
        if name:
            base_index_by_name[name] = i

    # Group the auditor block into (field column, optional following Diff column) pairs.
    # A column that can't be matched to a base field (even via alias) isn't necessarily a
    # mistake - some sheets carry auditor-only reference figures (e.g. a bond's Face Value
    # backing an Interest recalculation) that have no fund-side counterpart to check against.
    # Those are still shown alongside the base columns; they just don't get a correct/
    # incorrect verdict.
    validation_pairs = []  # (auditor_col_index, diff_col_index_or_None, base_field_name)
    unmatched_cols: list[int] = []
    k = 0
    while k < len(auditor_cols):
        col = auditor_cols[k]
        if _is_diff_header(row2[col]):
            k += 1  # stray Diff with nothing before it in the block
            continue
        diff_col = None
        if k + 1 < len(auditor_cols):
            next_col = auditor_cols[k + 1]
            if _is_diff_header(row2[next_col]):
                diff_col = next_col
                k += 1
        auditor_name = _normalize_header(row2[col])
        base_name = field_aliases.get(auditor_name, auditor_name)
        if base_name not in base_index_by_name:
            unmatched_cols.append(col)
        else:
            validation_pairs.append((col, diff_col, base_name))
        k += 1

    warnings: list[str] = []
    if unmatched_cols:
        unmatched_names = [row2[c] or f"column {c + 1}" for c in unmatched_cols]
        warnings.append(
            "Auditor column(s) with no matching original field, shown without validation: "
            + ", ".join(unmatched_names)
        )

    display_cols = sorted(base_cols + unmatched_cols)

    docs: list[dict] = []
    for r in range(3, max_row + 1):
        raw_values = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        if not any(str(v).strip() for v in raw_values if v is not None):
            continue

        row: dict[str, str] = {}
        for i in display_cols:
            header = row2[i]
            if not header:
                continue
            row[header] = "" if raw_values[i] is None else str(raw_values[i]).strip()

        if not row:
            warnings.append(f"Row {r}: skipped - no recognizable data")
            continue

        errors = []
        for auditor_col, diff_col, base_name in validation_pairs:
            base_i = base_index_by_name[base_name]
            original_value = row.get(row2[base_i], "")
            auditor_raw = raw_values[auditor_col]
            auditor_value = "" if auditor_raw is None else str(auditor_raw).strip()

            if diff_col is not None:
                diff_raw = raw_values[diff_col]
                if isinstance(diff_raw, bool):
                    # Some sheets precompute the Diff column as a match/no-match boolean
                    # (True = matches) rather than a numeric difference.
                    diff_value = None
                    mismatch = not diff_raw
                else:
                    diff_value = _to_float_or_none(diff_raw)
                    mismatch = diff_value is not None and abs(diff_value) > DIFF_TOLERANCE
            else:
                diff_value = None
                original_num = _to_float_or_none(original_value)
                auditor_num = _to_float_or_none(auditor_value)
                if original_num is not None and auditor_num is not None:
                    mismatch = abs(original_num - auditor_num) > DIFF_TOLERANCE
                    diff_value = auditor_num - original_num
                else:
                    mismatch = original_value.strip() != auditor_value.strip()

            if mismatch:
                errors.append({
                    "field": row2[base_i],
                    "original": original_value,
                    "auditor": auditor_value,
                    "diff": diff_value,
                })

        docs.append({
            "data": row,
            "status": "incorrect" if errors else ("correct" if validation_pairs else None),
            "errors": errors,
        })

    return docs, warnings


def parse_realised_gain(filename: str, content: bytes) -> tuple[list[dict], list[str]]:
    return _parse_sheet_with_auditor_validation(filename, content, "Realised Gain")


# The auditor block on Unrealised Gain sheets re-checks Holding Qty / Market Rate /
# Market Value under the shorter names Quantity / Rate / Amount.
UNREALISED_GAIN_FIELD_ALIASES = {
    "quantity": "holdingqty",
    "rate": "marketrate",
    "amount": "marketvalue",
}


def parse_unrealised_gain(filename: str, content: bytes) -> tuple[list[dict], list[str]]:
    return _parse_sheet_with_auditor_validation(
        filename, content, "Unrealised Gain", field_aliases=UNREALISED_GAIN_FIELD_ALIASES
    )


# The auditor block on Corporate Action sheets re-checks the fund's Amount by recalculating
# it from a bond's Face Value; the recalculated figure itself is labeled "Interest" (that's
# the corporate action type most rows carry), so it's aliased to Amount for the cross-check.
# Face Value has no fund-side counterpart - it's just the input the auditor computed from -
# so it's left unaliased and comes through as a plain, unvalidated column (see the
# unmatched-column handling in _parse_sheet_with_auditor_validation).
CORPORATE_ACTION_FIELD_ALIASES = {
    "interest": "amount",
}


def parse_corporate_action(filename: str, content: bytes) -> tuple[list[dict], list[str]]:
    return _parse_sheet_with_auditor_validation(
        filename, content, "Corporate Action", field_aliases=CORPORATE_ACTION_FIELD_ALIASES
    )


# The auditor block on Management Fees sheets re-checks Class Code / Fee% / Investor Code
# under the shorter/renamed labels Class / Rate / "NTPF as per JHS" (JHS being the auditor)
# respectively - confirmed by comparing each auditor column's values against candidate base
# columns row-by-row. Each is followed by a boolean Diff column (True = matches) rather than
# a numeric one.
MANAGEMENT_FEES_FIELD_ALIASES = {
    "class": "classcode",
    "rate": "fee",
    "ntpfasperjhs": "investorcode",
}


def parse_management_fees(filename: str, content: bytes) -> tuple[list[dict], list[str]]:
    return _parse_sheet_with_auditor_validation(
        filename, content, "Management Fees", field_aliases=MANAGEMENT_FEES_FIELD_ALIASES
    )


# SOA Closing sheets carry the same two-row-header auditor block as the NAV gain
# categories above - a merged "As per Validator" label over a repeating (field, Diff)
# block re-checking Class / NAV/Unit / Units against the fund's own figures. Every
# auditor column here is named exactly like the base field it re-checks, so no aliases
# are needed.
def parse_soa_closing(filename: str, content: bytes) -> tuple[list[dict], list[str]]:
    return _parse_sheet_with_auditor_validation(filename, content, "Closing")


# SOA XIRR sheets carry the same auditor block too - three base columns (Client Code,
# NAV Date, SOA) plus a merged "As per validator" label over the auditor's recalculated
# Rate + Diff. The auditor column is named "Rate" rather than "SOA", so it needs an
# alias to be matched against the base SOA column it's re-checking.
XIRR_FIELD_ALIASES = {"rate": "soa"}


def parse_soa_xirr(filename: str, content: bytes) -> tuple[list[dict], list[str]]:
    return _parse_sheet_with_auditor_validation(filename, content, "XIRR", field_aliases=XIRR_FIELD_ALIASES)


def _format_month_label(value) -> str:
    if hasattr(value, "strftime"):
        return value.strftime("%b %Y")
    return str(value).strip()


def parse_other_expense(filename: str, content: bytes) -> tuple[list[dict], list[str]]:
    """Parses the "Other Expense" upload: one row per expense type (Broking Fee,
    Operating Exp, GST Expense, Stamp Duty, ...), with a repeating block of three
    columns per month (Exp by FA / Exp by Auditor / Diff). The set of expense types
    isn't fixed - whatever row labels the admin's sheet has become the categories
    shown in the app, grouped dynamically rather than against a fixed list.

    Only .xlsx/.xlsm carry the merged month headers this layout depends on; CSV and
    legacy .xls uploads fall back to the flat parse_generic() read.
    """
    lower = (filename or "").lower()
    if not lower.endswith((".xlsx", ".xlsm")):
        df = read_table(filename, content)
        return parse_generic(df)

    try:
        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(content), data_only=True)
        ws = wb.active
    except Exception as exc:
        raise ValueError(f"Could not read '{filename}' as an Excel file: {exc}")

    max_col = ws.max_column
    max_row = ws.max_row
    if max_row < 3:
        raise ValueError("No data rows found in the uploaded file for Other Expense.")

    # Each month is a merged header cell spanning its 3 sub-columns (Exp by FA / Exp by
    # Auditor / Diff). Row-direction merges (the row-label column header, e.g. B1:B2)
    # span multiple rows in one column instead, so they're excluded by the max_col >
    # min_col check.
    month_blocks = []  # (start_col_0idx, end_col_0idx_exclusive, label)
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row == 1 and merged_range.max_row == 1 and merged_range.max_col > merged_range.min_col:
            label_value = ws.cell(row=1, column=merged_range.min_col).value
            month_blocks.append((merged_range.min_col - 1, merged_range.max_col, label_value))
    month_blocks.sort(key=lambda b: b[0])

    if not month_blocks:
        raise ValueError("Could not find any month columns in the uploaded Other Expense file.")

    # The expense-type label sits in the column immediately left of the first month block.
    label_col = month_blocks[0][0] - 1
    if label_col < 0:
        raise ValueError("Could not find the expense type column in the uploaded Other Expense file.")

    row2 = [str(ws.cell(row=2, column=c).value or "").strip() for c in range(1, max_col + 1)]

    docs: list[dict] = []
    warnings: list[str] = []

    for r in range(3, max_row + 1):
        expense_type = str(ws.cell(row=r, column=label_col + 1).value or "").strip()
        if not expense_type:
            continue

        for start, end, month_value in month_blocks:
            month_label = _format_month_label(month_value)
            block_headers = {_normalize_header(row2[c]): c for c in range(start, end)}
            fa_col = block_headers.get("expbyfa")
            auditor_col = block_headers.get("expbyauditor")
            diff_col = next((c for h, c in block_headers.items() if h in ("diff", "difference")), None)
            if fa_col is None:
                continue

            fa_value = ws.cell(row=r, column=fa_col + 1).value
            auditor_value = ws.cell(row=r, column=auditor_col + 1).value if auditor_col is not None else None
            if fa_value is None and auditor_value is None:
                continue

            status = None
            errors = []
            if auditor_col is not None:
                diff_num = _to_float_or_none(ws.cell(row=r, column=diff_col + 1).value) if diff_col is not None else None
                if diff_num is None:
                    fa_num = _to_float_or_none(fa_value)
                    auditor_num = _to_float_or_none(auditor_value)
                    diff_num = auditor_num - fa_num if fa_num is not None and auditor_num is not None else None
                mismatch = diff_num is not None and abs(diff_num) > DIFF_TOLERANCE
                status = "incorrect" if mismatch else "correct"
                if mismatch:
                    errors.append({
                        "field": "Amount",
                        "original": "" if fa_value is None else str(fa_value).strip(),
                        "auditor": "" if auditor_value is None else str(auditor_value).strip(),
                        "diff": diff_num,
                    })

            docs.append({
                "data": {
                    "Expense Type": expense_type,
                    "Month": month_label,
                    "Amount": "" if fa_value is None else str(fa_value).strip(),
                },
                "status": status,
                "errors": errors,
            })

    if not docs:
        warnings.append("No expense rows found in the uploaded file.")

    return docs, warnings


# Header name variants seen across different admins' Corpus In/Out templates over time.
# Uploaders aren't on one fixed template, so matching is deliberately generous here -
# see _resolve_investor_identity for how a row still gets kept even if only one of
# Investor Code / Investor Name is recognized.
_INVESTOR_CODE_ALIASES = (
    "investorcode", "clientcode", "clientid", "code", "invcode",
    "investorid", "clientnumber", "accountcode", "pan",
)
_INVESTOR_NAME_ALIASES = (
    "investorname", "clientname", "name", "investor", "nameofinvestor", "holdername",
)


def _resolve_investor_identity(norm: dict) -> tuple[str, str]:
    """Returns (investor_code, investor_name), falling back to whichever of the two is
    present when the other's column wasn't recognized - a row with a real identifier
    shouldn't be dropped just because its header for the *other* field used unfamiliar
    wording, since the uploaded Excel format varies between admins.
    """
    investor_code = _text(norm, *_INVESTOR_CODE_ALIASES, default="")
    investor_name = _text(norm, *_INVESTOR_NAME_ALIASES, default="")
    if not investor_code:
        investor_code = investor_name
    if not investor_name:
        investor_name = investor_code
    return investor_code, investor_name


# Corpus In's auditor block re-checks Allotment Type and Unit as per client under
# renamed labels ("...(Topup/Fresh)" / "Unit Allocated"); Class code is named identically
# on both sides so it needs no alias.
CORPUS_IN_FIELD_ALIASES = {
    "allotmenttypetopupfresh": "allotmenttypeasperclient",
    "unitallocated": "unitasperclient",
}


def parse_corpus_in(filename: str, content: bytes) -> tuple[list[dict], list[str]]:
    raw_docs, warnings = _parse_sheet_with_auditor_validation(
        filename, content, "Corpus In", field_aliases=CORPUS_IN_FIELD_ALIASES
    )
    docs: list[dict] = []

    for i, raw in enumerate(raw_docs, start=2):
        row = raw["data"]
        norm = {_normalize_header(k): v for k, v in row.items()}
        investor_code, investor_name = _resolve_investor_identity(norm)
        if not investor_code and not investor_name:
            warnings.append(f"Row {i}: skipped - missing Investor Code and Investor Name")
            continue

        amount = _number(norm, "capitalreceived", "amount", "contributionamount", "amountreceived", "capitalamount")

        docs.append({
            "movement_type": "In",
            "movement_date": _text(norm, "dateofbank", "date", "transactiondate", "valuedate", "dateofcontribution"),
            "investor_code": investor_code,
            "investor_name": investor_name,
            "client_class": _text(norm, "classcode", "class", "clientclass"),
            "bank_account": "",
            "amount": amount,
            "data": row,
            "status": raw["status"],
            "errors": raw["errors"],
        })

    return docs, warnings


# Corpus Out's auditor block re-checks Class under the renamed label "Updated Class";
# Bank Name / Bank A/C No / IFSC are named identically on both sides so need no alias.
# Waiveoff Approval and Remarks have no fund-side counterpart to re-check - they come
# through as plain unmatched (unvalidated) columns instead.
CORPUS_OUT_FIELD_ALIASES = {
    "updatedclass": "class",
}


def parse_corpus_out(filename: str, content: bytes) -> tuple[list[dict], list[str]]:
    raw_docs, warnings = _parse_sheet_with_auditor_validation(
        filename, content, "Corpus Out", field_aliases=CORPUS_OUT_FIELD_ALIASES
    )
    docs: list[dict] = []

    for i, raw in enumerate(raw_docs, start=2):
        row = raw["data"]
        norm = {_normalize_header(k): v for k, v in row.items()}
        investor_code, investor_name = _resolve_investor_identity(norm)
        if not investor_code and not investor_name:
            warnings.append(f"Row {i}: skipped - missing Investor Code and Investor Name")
            continue

        # There's no single "Amount" column - the actual redeemed amount is what's left
        # after "Capital Pending" (not yet processed) is taken out of "Capital As On".
        # Many sheets also carry an "As per Email" column with the confirmed redemption
        # figure; prefer that when present. Falls back to a direct redemption-amount
        # column when the sheet has one instead.
        if _text(norm, "asperemail", default=""):
            amount = _number(norm, "asperemail")
        elif _text(norm, "capitalason", "capitalbalance", "balanceason", default=""):
            amount = _number(norm, "capitalason", "capitalbalance", "balanceason") - _number(
                norm, "capitalpending", "pendingamount", "pending"
            )
        else:
            amount = _number(norm, "amount", "redemptionamount", "capitalredeemed")

        docs.append({
            "movement_type": "Out",
            "movement_date": _text(norm, "redemptionrequestdate", "dateofbank", "date", "transactiondate", "valuedate"),
            "investor_code": investor_code,
            "investor_name": investor_name,
            "client_class": _text(norm, "class", "classcode", "clientclass"),
            "bank_account": _text(norm, "bankacno", "bankaccountno", "bankaccount", "accountno", "accountnumber"),
            "amount": amount,
            "data": row,
            "status": raw["status"],
            "errors": raw["errors"],
        })

    return docs, warnings


# Dashboard upload: one Excel workbook with three fixed-name sheets, each feeding a
# different part of the fund's Dashboard page.
DASHBOARD_SHEET_CATEGORIES = {
    "Fund NAV": "fund-nav",
    "XIRR": "xirr",
    "Client Master": "client-master",
}


def _read_dashboard_sheet(ws) -> list[dict]:
    """Reads one Dashboard sheet, keyed by its own header text (not normalized).

    Unlike the other multi-row-header sheets in this file, the header row here isn't
    reliably identified by a merge - row 1 sometimes carries a stray leftover value or
    two (e.g. a lone number) instead of real column labels, with the actual headers one
    row down. Whichever of row 1 / row 2 has more text-looking cells is treated as the
    real header row, since a genuine header row is almost entirely text labels while a
    row of stray leftovers isn't.
    """
    max_col = ws.max_column
    max_row = ws.max_row

    def text_cell_count(row_values: list) -> int:
        return sum(1 for v in row_values if isinstance(v, str) and v.strip())

    row1 = [ws.cell(row=1, column=c).value for c in range(1, max_col + 1)]
    row2 = [ws.cell(row=2, column=c).value for c in range(1, max_col + 1)] if max_row >= 2 else []
    header_row = 2 if text_cell_count(row2) > text_cell_count(row1) else 1
    data_start_row = header_row + 1

    if max_row < data_start_row:
        return []

    headers = [str(ws.cell(row=header_row, column=c).value or "").strip() for c in range(1, max_col + 1)]

    rows: list[dict] = []
    for r in range(data_start_row, max_row + 1):
        raw_values = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        if not any(str(v).strip() for v in raw_values if v is not None):
            continue
        row: dict[str, str] = {}
        for i, header in enumerate(headers):
            if not header:
                continue
            value = raw_values[i]
            row[header] = "" if value is None else str(value).strip()
        if row:
            rows.append(row)

    return rows


# Category prefix for sheets beyond the three fixed ones - any extra sheet in the
# workbook is picked up automatically (no admin prompt for what it is), stored under
# this prefix + the sheet's own name so it can't collide with the fixed categories, and
# charted with a type auto-guessed from its shape on the frontend (see
# detectAutoChartKind in frontend/js/app.js) rather than asked about at upload time.
EXTRA_SHEET_CATEGORY_PREFIX = "sheet:"


def parse_dashboard_workbook(filename: str, content: bytes) -> tuple[dict[str, list[dict]], list[str]]:
    """Parses the Dashboard upload's sheets into {category_slug: rows}. The three
    fixed sheets (Fund NAV / XIRR / Client Master, matched by exact name) map to their
    own known category; any other sheet in the workbook is picked up too, under
    "sheet:<its own name>", so an admin can add more charts to the Dashboard just by
    adding a sheet - no upload-time prompt for what it is. A missing fixed sheet is
    warned about and simply excluded rather than failing the whole upload, since an
    admin may upload a partial Dashboard file while waiting on the rest.
    """
    lower = (filename or "").lower()
    if not lower.endswith((".xlsx", ".xlsm")):
        raise ValueError(
            "The Dashboard upload must be an Excel file (.xlsx/.xlsm) with \"Fund NAV\", "
            "\"XIRR\", and \"Client Master\" sheets."
        )

    try:
        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(content), data_only=True)
    except Exception as exc:
        raise ValueError(f"Could not read '{filename}' as an Excel file: {exc}")

    results: dict[str, list[dict]] = {}
    warnings: list[str] = []
    for sheet_name, category in DASHBOARD_SHEET_CATEGORIES.items():
        if sheet_name not in wb.sheetnames:
            warnings.append(f"Sheet '{sheet_name}' not found in the uploaded file - skipped.")
            continue
        results[category] = _read_dashboard_sheet(wb[sheet_name])

    for sheet_name in wb.sheetnames:
        if sheet_name in DASHBOARD_SHEET_CATEGORIES:
            continue
        rows = _read_dashboard_sheet(wb[sheet_name])
        if rows:
            results[f"{EXTRA_SHEET_CATEGORY_PREFIX}{sheet_name}"] = rows

    if not any(results.values()):
        raise ValueError(
            "No data found in the expected sheets (\"Fund NAV\", \"XIRR\", \"Client Master\") "
            "of the uploaded file."
        )

    return results, warnings
